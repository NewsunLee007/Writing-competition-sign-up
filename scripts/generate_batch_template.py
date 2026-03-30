from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


root = Path(__file__).resolve().parents[1] / "frontend" / "public"
root.mkdir(parents=True, exist_ok=True)
file_path = root / "batch-registration-template.xlsx"

wb = Workbook()
guide_ws = wb.active
guide_ws.title = "填写说明"
dict_ws = wb.create_sheet("字典")
data_ws = wb.create_sheet("报名数据")

navy = "10203C"
cream = "F7F1E7"
line = "D9CFBF"
white = "FFFFFF"
header_fill = PatternFill("solid", fgColor=navy)
soft_fill = PatternFill("solid", fgColor=cream)
thin = Side(style="thin", color=line)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

guide_ws["A1"] = "瑞安市英语创意写作评审活动批量报名模板"
guide_ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=16, color=navy)
guide_ws["A3"] = "填写规则"
guide_ws["A3"].font = Font(name="Microsoft YaHei", bold=True, size=12, color=navy)

rules = [
    "1. 报名类别仅可选择“学区”或“直属学校”。",
    "2. “学区/直属学校”列已配置下拉列表。",
    "3. 若报名类别为直属学校，“学校”列会自动带出学校名称。",
    "4. 若报名类别为学区，请在“学校”列手动填写具体学校名称。",
    "5. 每位学生都必须填写指导教师、带队教师、带队教师电话。",
    "6. 建议导入前先另存模板，保留原始空白模板。",
]

for idx, text in enumerate(rules, start=4):
    guide_ws[f"A{idx}"] = text
    guide_ws[f"A{idx}"].alignment = Alignment(wrap_text=True)

guide_ws["A11"] = "字段说明"
guide_ws["A11"].font = Font(name="Microsoft YaHei", bold=True, size=12, color=navy)

field_rows = [
    ("报名类别", "下拉选择：学区 / 直属学校"),
    ("学区/直属学校", "下拉选择具体归属"),
    ("学校", "学区报名时手填，直属学校自动带出"),
    ("学生姓名", "必填"),
    ("指导教师", "必填，一人一对应"),
    ("带队教师", "必填"),
    ("带队教师电话", "必填，11位中国大陆手机号"),
]

for row_index, (name, desc) in enumerate(field_rows, start=12):
    guide_ws[f"A{row_index}"] = name
    guide_ws[f"B{row_index}"] = desc
    guide_ws[f"A{row_index}"].font = Font(name="Microsoft YaHei", bold=True, color=navy)
    guide_ws[f"A{row_index}"].fill = soft_fill
    guide_ws[f"B{row_index}"].fill = soft_fill
    guide_ws[f"A{row_index}"].border = border
    guide_ws[f"B{row_index}"].border = border

guide_ws.column_dimensions["A"].width = 24
guide_ws.column_dimensions["B"].width = 48

districts = ["塘下学区", "安阳学区", "飞云学区", "莘塍学区", "马屿学区", "高楼学区", "湖岭学区", "陶山学区"]
direct_schools = ["安阳实验", "新纪元", "安高初中", "瑞祥实验", "集云学校", "毓蒙中学", "广场中学", "瑞中附初", "紫荆书院"]

dict_ws["A1"] = "报名类别"
dict_ws["A2"] = "学区"
dict_ws["A3"] = "直属学校"

for index, value in enumerate(districts, start=1):
    dict_ws.cell(index, 2, value)

for index, value in enumerate(direct_schools, start=1):
    dict_ws.cell(index, 3, value)

wb.defined_names.add(DefinedName("TypeOptions", attr_text="字典!$A$2:$A$3"))
wb.defined_names.add(DefinedName("DistrictOptions", attr_text=f"字典!$B$1:$B${len(districts)}"))
wb.defined_names.add(DefinedName("DirectSchoolOptions", attr_text=f"字典!$C$1:$C${len(direct_schools)}"))

headers = ["报名类别", "学区/直属学校", "学校", "学生姓名", "指导教师", "带队教师", "带队教师电话"]

for col_index, header in enumerate(headers, start=1):
    cell = data_ws.cell(1, col_index, header)
    cell.font = Font(name="Microsoft YaHei", bold=True, color=white)
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

for idx, width in enumerate([16, 24, 24, 16, 16, 16, 18], start=1):
    data_ws.column_dimensions[chr(64 + idx)].width = width

data_ws.freeze_panes = "A2"

type_validation = DataValidation(type="list", formula1="=TypeOptions", allow_blank=False)
unit_validation = DataValidation(
    type="list",
    formula1='=INDIRECT(IF($A2="直属学校","DirectSchoolOptions","DistrictOptions"))',
    allow_blank=False,
)

for validation in [type_validation, unit_validation]:
    validation.promptTitle = "下拉选择"
    validation.prompt = "请通过下拉列表选择有效值"
    validation.errorTitle = "输入无效"
    validation.error = "该单元格只能选择模板下拉中的值"
    data_ws.add_data_validation(validation)

type_validation.add("A2:A300")
unit_validation.add("B2:B300")

for row in range(2, 301):
    data_ws[f"C{row}"] = f'=IF($A{row}="直属学校",$B{row},"")'
    for col in range(1, 8):
        data_ws.cell(row, col).border = border
        if row % 2 == 0:
            data_ws.cell(row, col).fill = PatternFill("solid", fgColor="FCF8F1")

guide_ws.sheet_view.showGridLines = False
data_ws.sheet_view.showGridLines = False
dict_ws.sheet_state = "hidden"

wb.save(file_path)
print(file_path)
